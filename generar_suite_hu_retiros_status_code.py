import openpyxl
from datetime import datetime

INPUT_PATH = "Hu Retiros.xlsx"
TEMPLATE_PATH = "casos de prueba/plantilla_base.xlsx"
OUTPUT_PATH = "casos de prueba/hu_retiros_status_code_suite.xlsx"

# Códigos permitidos (del contexto del proyecto)
ALLOWED_STATUS_CODES = {200, 204, 100, 300, 600, 700, 900, 901}

COLS = [
    "Issue ID",
    "Tipo de test",
    "Resumen",
    "Descripcion",
    "Escenario",
    "Resultado Final",
    "Accion",
    "Datos",
    "Resultado Esperado",
]


def _clear_rows_keep_header(ws):
    for r in range(2, ws.max_row + 1):
        for c in range(1, 10):
            ws.cell(row=r, column=c).value = None


def _ensure_headers(ws):
    current = [ws.cell(row=1, column=i).value for i in range(1, 10)]
    if any(v is None for v in current):
        for i, col in enumerate(COLS, start=1):
            ws.cell(row=1, column=i).value = col


def _parse_expected_http_codes(salidas_esperadas: str, escenarios_pos: str, escenarios_neg: str) -> list[int]:
    """
    Extrae códigos HTTP desde texto tipo: 'HTTP 204 esperado' o 'HTTP 201; X-RqUID...'
    y también soporta textos tipo: 'StatusCode 200' o 'código 200' (propio de HUs de retiro).
    Regla: SOLO se aceptan códigos en ALLOWED_STATUS_CODES.
    """
    codes: set[int] = set()

    def scan(src: str):
        if not src:
            return
        parts = src.replace(";", " ").replace(",", " ").replace(":", " ").split()
        upper = [p.upper() for p in parts]

        # Patrón 1: HTTP <num>
        for i, token in enumerate(upper):
            if token == "HTTP" and i + 1 < len(parts) and parts[i + 1].strip().isdigit():
                code = int(parts[i + 1].strip())
                if code in ALLOWED_STATUS_CODES:
                    codes.add(code)

        # Patrón 2: STATUSCODE <num>
        for i, token in enumerate(upper):
            if token == "STATUSCODE" and i + 1 < len(parts) and parts[i + 1].strip().isdigit():
                code = int(parts[i + 1].strip())
                if code in ALLOWED_STATUS_CODES:
                    codes.add(code)

        # Patrón 3: CODIGO / CÓDIGO <num>
        for i, token in enumerate(upper):
            if token in ("CODIGO", "CÓDIGO") and i + 1 < len(parts) and parts[i + 1].strip().isdigit():
                code = int(parts[i + 1].strip())
                if code in ALLOWED_STATUS_CODES:
                    codes.add(code)

    for src in (salidas_esperadas, escenarios_pos, escenarios_neg):
        scan(src)

    return sorted(codes)


def generar():
    # Lee input HUs
    wb_in = openpyxl.load_workbook(INPUT_PATH)
    ws_in = wb_in.active

    header = [c.value for c in ws_in[1]]
    idx = {name: header.index(name) for name in header if name is not None}

    def get(row, col_name):
        pos = idx.get(col_name)
        return row[pos].value if pos is not None else None

    casos = []
    next_id = 1

    for row in ws_in.iter_rows(min_row=2):
        hu_id = get(row, "ID_HU")
        modulo = get(row, "Módulo")
        nombre = get(row, "Nombre_HU")
        api = get(row, "API_Relacionada")
        method = get(row, "Método_HTTP")
        salidas = get(row, "Salidas_Esperadas")
        esc_pos = get(row, "Escenarios_Positivos")
        esc_neg = get(row, "Escenarios_Negativos")

        codes = _parse_expected_http_codes(str(salidas or ""), str(esc_pos or ""), str(esc_neg or ""))

        # Si una HU no trae explícito un código permitido, no se inventa.
        if not codes:
            continue

        for code in codes:
            casos.append(
                [
                    next_id,
                    "Automatizado",
                    f"[{hu_id}] {modulo} - {nombre} - Validar HTTP {code}",
                    f"Validar únicamente que el endpoint responde con status code HTTP {code} (según códigos del contexto del proyecto).",
                    "Se dispone de respuesta o mock que retorna el status code esperado.",
                    "Pending",
                    f"{method} {api} (mock/condición que retorna {code}).",
                    "",
                    f"HTTP {code}",
                ]
            )
            next_id += 1

    # Carga plantilla y escribe salida
    wb_out = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_out = wb_out.active

    _ensure_headers(ws_out)
    _clear_rows_keep_header(ws_out)

    for r, caso in enumerate(casos, start=2):
        for c, val in enumerate(caso, start=1):
            ws_out.cell(row=r, column=c, value=val)

    wb_out.save(OUTPUT_PATH)
    print(f"Generados {len(casos)} casos en {OUTPUT_PATH}")


if __name__ == "__main__":
    generar()
