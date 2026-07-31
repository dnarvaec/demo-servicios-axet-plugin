"""
jira_uploader.py
────────────────────────────────────────────────────────────────────────────
Sube los casos de prueba de un Excel generado al proyecto Jira configurado
en el archivo .env de la raíz del proyecto.

Uso directo:
    python jira_uploader.py casos\ de\ prueba/retiro_otp.xlsx

Uso desde otro script:
    from jira_uploader import subir_casos_a_jira
    resultados = subir_casos_a_jira("casos de prueba/retiro_otp.xlsx")
"""

import base64
import os
import sys
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

# Cargar .env desde la raíz del proyecto (donde vive este script)
load_dotenv(Path(__file__).parent / ".env")

# ─── Configuración desde .env ────────────────────────────────────────────────
JIRA_BASE_URL: str = os.getenv("JIRA_BASE_URL", "").rstrip("/")
JIRA_EMAIL: str = os.getenv("JIRA_EMAIL", "")
JIRA_API_TOKEN: str = os.getenv("JIRA_API_TOKEN", "")
JIRA_PROJECT_KEY: str = os.getenv("JIRA_PROJECT_KEY", "")
JIRA_ISSUE_TYPE: str = os.getenv("JIRA_ISSUE_TYPE", "Test")
JIRA_API_VERSION: str = os.getenv("JIRA_API_VERSION", "3")
JIRA_AUTH_TYPE: str = os.getenv("JIRA_AUTH_TYPE", "basic").strip().lower()
JIRA_UPLOAD_ENABLED: bool = os.getenv("JIRA_UPLOAD_ENABLED", "true").strip().lower() == "true"

# Custom fields opcionales
JIRA_FIELD_PRECONDICION: str = os.getenv("JIRA_FIELD_PRECONDICION", "")
JIRA_FIELD_PASOS: str = os.getenv("JIRA_FIELD_PASOS", "")
JIRA_FIELD_DATOS: str = os.getenv("JIRA_FIELD_DATOS", "")
JIRA_FIELD_RESULTADO_ESPERADO: str = os.getenv("JIRA_FIELD_RESULTADO_ESPERADO", "")

# Columnas esperadas en el Excel (deben coincidir con la plantilla)
EXCEL_COLS = [
    "Issue ID",
    "Tipo de test",
    "Resumen",
    "Descripcion",
    "Escenario",
    "Resultado Final",
    "Accion",
    "Datos",
    "Resultado Esperado",
]

# Nombre de la columna que se agrega al Excel con la clave Jira creada
JIRA_KEY_COL = "Jira Key"


# ─── Autenticación ───────────────────────────────────────────────────────────

def _auth_headers() -> dict:
    # Bearer = PAT de Jira Server/DC (≥ 8.14); basic = usuario:contraseña
    if JIRA_AUTH_TYPE == "bearer":
        return {
            "Authorization": f"Bearer {JIRA_API_TOKEN}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
    credentials = base64.b64encode(
        f"{JIRA_EMAIL}:{JIRA_API_TOKEN}".encode()
    ).decode()
    return {
        "Authorization": f"Basic {credentials}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


# ─── Construcción del payload ────────────────────────────────────────────────

def _text(value) -> str:
    return str(value).strip() if value is not None else "—"


def _adf_heading(text: str, level: int = 3) -> dict:
    return {
        "type": "heading",
        "attrs": {"level": level},
        "content": [{"type": "text", "text": text}],
    }


def _adf_paragraph(text: str) -> dict:
    return {
        "type": "paragraph",
        "content": [{"type": "text", "text": _text(text)}],
    }


def _build_description_adf(descripcion, escenario, accion, datos, resultado_esperado) -> dict:
    """Construye la descripción en Atlassian Document Format (API v3)."""
    content = [_adf_heading("Descripción", 3), _adf_paragraph(descripcion)]

    if escenario:
        content += [_adf_heading("Precondición / Escenario", 3), _adf_paragraph(escenario)]
    if accion:
        content += [_adf_heading("Acción / Pasos", 3), _adf_paragraph(accion)]
    if datos:
        content += [_adf_heading("Datos de Prueba", 3), _adf_paragraph(datos)]
    if resultado_esperado:
        content += [_adf_heading("Resultado Esperado", 3), _adf_paragraph(resultado_esperado)]

    return {"type": "doc", "version": 1, "content": content}


def _build_description_v2(descripcion, escenario, accion, datos, resultado_esperado) -> str:
    """Construye la descripción en texto plano con wiki markup (API v2)."""
    parts = [f"h3. Descripción\n{_text(descripcion)}"]
    if escenario:
        parts.append(f"h3. Precondición / Escenario\n{_text(escenario)}")
    if accion:
        parts.append(f"h3. Acción / Pasos\n{_text(accion)}")
    if datos:
        parts.append(f"h3. Datos de Prueba\n{_text(datos)}")
    if resultado_esperado:
        parts.append(f"h3. Resultado Esperado\n{_text(resultado_esperado)}")
    return "\n\n".join(parts)


def _build_payload(row: dict) -> dict:
    issue_id = row.get("Issue ID")
    resumen = _text(row.get("Resumen"))
    tipo = _text(row.get("Tipo de test"))
    descripcion = row.get("Descripcion")
    escenario = row.get("Escenario")
    accion = row.get("Accion")
    datos = row.get("Datos")
    resultado_esperado = row.get("Resultado Esperado")

    summary = f"[{issue_id}] {resumen}" if issue_id else resumen

    # Descripción: si hay custom fields, el cuerpo solo lleva la descripción base
    use_custom_fields = bool(JIRA_FIELD_PRECONDICION)

    if JIRA_API_VERSION == "3":
        if use_custom_fields:
            description = _build_description_adf(descripcion, None, None, None, None)
        else:
            description = _build_description_adf(descripcion, escenario, accion, datos, resultado_esperado)
    else:
        if use_custom_fields:
            description = _build_description_v2(descripcion, None, None, None, None)
        else:
            description = _build_description_v2(descripcion, escenario, accion, datos, resultado_esperado)

    label = tipo.replace(" ", "_") if tipo and tipo != "—" else None

    fields: dict = {
        "project": {"key": JIRA_PROJECT_KEY},
        "summary": summary,
        "issuetype": {"name": JIRA_ISSUE_TYPE},
        "description": description,
    }

    if label:
        fields["labels"] = [label]

    if use_custom_fields:
        if JIRA_FIELD_PRECONDICION and escenario:
            fields[JIRA_FIELD_PRECONDICION] = _text(escenario)
        if JIRA_FIELD_PASOS and accion:
            fields[JIRA_FIELD_PASOS] = _text(accion)
        if JIRA_FIELD_DATOS and datos:
            fields[JIRA_FIELD_DATOS] = _text(datos)
        if JIRA_FIELD_RESULTADO_ESPERADO and resultado_esperado:
            fields[JIRA_FIELD_RESULTADO_ESPERADO] = _text(resultado_esperado)

    return {"fields": fields}


# ─── Llamada a la API ─────────────────────────────────────────────────────────

def _crear_issue(row: dict) -> str | None:
    """Crea un issue en Jira y devuelve la clave (p.ej. EV-42) o None si falla."""
    url = f"{JIRA_BASE_URL}/rest/api/{JIRA_API_VERSION}/issue"
    payload = _build_payload(row)

    try:
        resp = requests.post(url, headers=_auth_headers(), json=payload, timeout=30)
    except requests.RequestException as exc:
        print(f"  ✗ Error de conexión: {exc}")
        return None

    if resp.status_code == 201:
        return resp.json().get("key")

    print(f"  ✗ HTTP {resp.status_code} — {resp.text[:300]}")
    return None


# ─── Función principal ────────────────────────────────────────────────────────

def subir_casos_a_jira(excel_path: str) -> dict[str, str]:
    """
    Lee el Excel y sube cada caso de prueba como un issue Jira.

    Devuelve un dict {str(issue_id): jira_key} con los issues creados.
    También añade la columna «Jira Key» al Excel y lo guarda.
    """
    if not JIRA_UPLOAD_ENABLED:
        print("JIRA_UPLOAD_ENABLED=false — subida a Jira omitida.")
        return {}

    missing = [v for v in ("JIRA_BASE_URL", "JIRA_API_TOKEN", "JIRA_PROJECT_KEY")
               if not os.getenv(v)]
    if JIRA_AUTH_TYPE == "basic" and not os.getenv("JIRA_EMAIL"):
        missing.append("JIRA_EMAIL")
    if missing:
        raise EnvironmentError(
            f"Faltan variables de entorno en .env: {', '.join(missing)}"
        )

    ruta = Path(excel_path)
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {ruta.resolve()}")

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active

    # Leer encabezados de la fila 1
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Añadir columna "Jira Key" si no existe
    if JIRA_KEY_COL not in headers:
        jira_key_col_idx = len(headers) + 1
        ws.cell(row=1, column=jira_key_col_idx, value=JIRA_KEY_COL)
    else:
        jira_key_col_idx = headers.index(JIRA_KEY_COL) + 1

    resultados: dict[str, str] = {}
    creados = 0
    errores = 0

    print(f"\nSubiendo a Jira → {JIRA_BASE_URL}  |  Proyecto: {JIRA_PROJECT_KEY}\n")

    for row_idx in range(2, ws.max_row + 1):
        row_data = {
            headers[c - 1]: ws.cell(row=row_idx, column=c).value
            for c in range(1, len(headers) + 1)
            if headers[c - 1] is not None
        }

        if not row_data.get("Resumen"):
            continue

        issue_id = row_data.get("Issue ID", row_idx - 1)
        resumen_corto = _text(row_data.get("Resumen"))[:55]
        print(f"  [{issue_id:>3}] {resumen_corto:<55}", end=" → ")

        jira_key = _crear_issue(row_data)
        if jira_key:
            resultados[str(issue_id)] = jira_key
            ws.cell(row=row_idx, column=jira_key_col_idx, value=jira_key)
            print(f"✓  {jira_key}")
            creados += 1
        else:
            errores += 1

    wb.save(ruta)

    print(f"\n{'─'*60}")
    print(f"  Creados : {creados}")
    print(f"  Errores : {errores}")
    print(f"  Excel   : {ruta.resolve()}")
    print(f"{'─'*60}\n")

    return resultados


# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python jira_uploader.py <ruta_excel>")
        print("Ejemplo: python jira_uploader.py \"casos de prueba/retiro_otp.xlsx\"")
        sys.exit(1)

    subir_casos_a_jira(sys.argv[1])
